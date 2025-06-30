import asyncio
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from humanLikeScroll import human_like_scroll
from datetime import datetime, timedelta
import re

async def eachMachineFunc(page, model_name):
    await asyncio.sleep(1)
    file_name = "result.xlsx"
    sheet_name = "パチンコ"

    # ファイルが存在するか確認
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        # ヘッダを書き込み
        headers = [
            "日付", "機種名", "台番号", "打込み玉数", "差玉", 
            "大当たり回数", "継続回数", "累計スタート", "最終スタート"
        ]
        ws.append(headers)


    # await page.wait_for_load_state("load")
    title = await page.title()
    print("New page title:", title)
    await human_like_scroll(page, 500)
    # await asyncio.sleep(5)

    extracted_data = [
        "2025-06-28",           # 日付
        title,                 # 機種名（ここではタイトルを機種名の代わりに使用）
        "123",                 # 台番号
        "1000",                # 打込み玉数
        "-200",                # 差玉
        "5",                   # 大当たり回数
        "2",                   # 継続回数
        "500",                 # 累計スタート
        "150",                 # 最終スタート
    ]
    
    # 現在の日付と時刻を取得
    now = datetime.now()

    machine_no = ''

    # div#divDAI の中の <h2> を取得
    h2_element = await page.query_selector('#divDAI h2')

    if h2_element:
        h2_text = await h2_element.inner_text()

        # 正規表現で最後の4桁の数字を抽出
        match = re.search(r'(\d{4})\s*$', h2_text)
        if match:
            machine_no = match.group(1)
        else:
            print("4桁の番号が見つかりませんでした。")
    else:
        print("div#divDAI h2 が見つかりませんでした。")

    # "scroll" クラスの div 内の <tr> を取得
    try:
        scroll_div = await page.wait_for_selector('div.scroll', timeout=3000)
        tr = await scroll_div.query_selector('tr')
    except Exception as e:
        print("⚠️ scroll_div または tr が見つかりませんでした:", e)
        return


    # tr 内のすべての <td> を取得
    tds = await tr.query_selector_all('td')

    result = []

    # 最初の td をスキップ（index 0）、2番目以降に対して処理
    for i, td in enumerate(tds[1:], start=1):
        if i > 6:
            break  # 最大6日前までに制限
        # td 内のすべての <div class="outer border-bottom"> を取得
        divs = await td.query_selector_all('div.outer.border-bottom')

        # index: 1, 2, 4, 5（0-based index）を抽出
        indices_to_extract = [1, 2, 4, 5]

        values = []
        for j in indices_to_extract:
            if j < len(divs):
                inner_div = await divs[j].query_selector('div.inner.nc-text-align-right')
                if inner_div:
                    text = await inner_div.inner_text()
                    values.append(text.strip())
                else:
                    values.append("")  # inner div not found
            else:
                values.append("")  # index out of range
        
        target_date = now - timedelta(days=i)
        formatted_date = target_date.strftime("%Y/%m/%d")
        extracted_data[0] = formatted_date
        extracted_data[1] = model_name
        extracted_data[2] = machine_no
        extracted_data[3] = extracted_data[4] = ''
        extracted_data[5] = values[0]
        extracted_data[6] = values[1]
        extracted_data[7] = values[2]
        extracted_data[8] = values[3]

        print(extracted_data)

        # データを書き込み
        ws.append(extracted_data)

        # 保存
        wb.save(file_name)
        print(f"✔️ Excelに保存完了（{file_name} - {sheet_name}）")